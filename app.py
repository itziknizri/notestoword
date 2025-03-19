import tkinter as tk
from tkinter import filedialog, ttk, messagebox, font
from docx import Document
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
import re
import os
from datetime import datetime
import tempfile

# ניסיון לייבא win32com - עשוי להיכשל אם החבילה לא מותקנת
try:
    import win32com.client
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

class WordCommentsExtractor:
    def __init__(self, root):
        self.root = root
        self.root.title("ממיר הערות מוורד לאקסל")
        self.root.geometry("1100x650")
        
        # ערכים שנשמור
        self.docx_path = None
        self.comments_data = []
        
        # בדיקה האם win32com זמין
        self.has_win32com = HAS_WIN32COM
        
        # הגדרת כיוון RTL
        self.configure_rtl_support()
        
        # יצירת ממשק
        self.create_widgets()
    
    def configure_rtl_support(self):
        """הגדרת תמיכה בכיוון מימין לשמאל"""
        # הגדרת פונטים
        default_font = font.nametofont("TkDefaultFont")
        default_font.configure(family="Segoe UI", size=10)
        
        # ניסיון להגדיר RTL באופן מערכתי
        try:
            self.root.tk.call('encoding', 'system', 'utf-8')
            self.root.tk_strictMotif(False)
        except:
            pass
        
        # הגדרת סגנון טקסט לימין
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), anchor="e")  # כותרות לימין
        style.configure("Treeview", font=("Segoe UI", 10), anchor="e")  # תאים לימין
        style.configure("TLabel", font=("Segoe UI", 10), anchor="e")  # טקסט לימין
        style.configure("TButton", font=("Segoe UI", 10), anchor="center")  # כפתורים מרכז
        
        # הגדרת צבעים
        self.root.configure(bg="#f0f0f0")
    
    def create_widgets(self):
        # מסגרת ראשית
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill="both", expand=True)
        
        # כותרת
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=10)
        
        header = ttk.Label(header_frame, text="ממיר הערות מקובץ וורד לאקסל", 
                          font=("Segoe UI", 16, "bold"))
        header.pack(side="right", padx=10)
        
        # מסגרת לבחירת קובץ (מסודרת מימין לשמאל)
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill="x", pady=5)
        
        select_btn = ttk.Button(file_frame, text="בחר קובץ וורד", command=self.select_file)
        select_btn.pack(side="right", padx=5)
        
        self.file_label = ttk.Label(file_frame, text="לא נבחר קובץ")
        self.file_label.pack(side="right", padx=5, fill="x", expand=True)
        
        # מסגרת כפתורי פעולה
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill="x", pady=5)
        
        # כפתורי פעולה בצד ימין
        process_btn = ttk.Button(action_frame, text="עבד את הקובץ", command=self.process_file)
        process_btn.pack(side="right", padx=5)
        
        export_btn = ttk.Button(action_frame, text="ייצא לאקסל", command=self.export_to_excel)
        export_btn.pack(side="right", padx=5)
        
        # כפתורי מיון - מוסיף אפשרויות מיון
        sort_label = ttk.Label(action_frame, text="מיון לפי:")
        sort_label.pack(side="right", padx=5)
        
        sort_page_btn = ttk.Button(action_frame, text="עמוד", 
                                  command=lambda: self.sort_treeview_by_column(1))
        sort_page_btn.pack(side="right", padx=2)
        
        sort_date_btn = ttk.Button(action_frame, text="תאריך", 
                                  command=lambda: self.sort_treeview_by_column(4))
        sort_date_btn.pack(side="right", padx=2)
        
        # אזור סטטוס
        status_frame = ttk.Frame(main_frame, relief="sunken", borderwidth=1)
        status_frame.pack(fill="x", pady=5)
        
        self.status_var = tk.StringVar()
        self.status_var.set("מוכן")
        status_label = ttk.Label(status_frame, textvariable=self.status_var)
        status_label.pack(side="right", pady=5, fill="x", expand=True)
        
        # יצירת מסגרת לטבלה
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True, pady=5)
        
        # הגדרת עמודות לפי הדרישות - סדר מימין לשמאל
        columns = (
            "מס'", "עמוד", "הערה", "כותב ההערה", "תאריך ההערה", 
            "תגובה 1", "כותב תגובה 1", "תאריך תגובה 1",
            "תגובה 2", "כותב תגובה 2", "תאריך תגובה 2",
            "תגובה 3", "כותב תגובה 3", "תאריך תגובה 3"
        )
        
        # יצירת טבלה עם סדר עמודות מימין לשמאל
        self.result_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=20)
        
        # הגדרת פונקציות מיון על ידי לחיצה על כותרות
        for i, col in enumerate(columns):
            self.result_tree.heading(col, text=col, anchor="e",
                                   command=lambda c=i: self.sort_treeview_by_column(c))
            
            # רוחב קבוע לעמודות
            if col == "מס'":
                width = 40
            elif col == "עמוד":
                width = 50
            elif col in ["תאריך ההערה", "תאריך תגובה 1", "תאריך תגובה 2", "תאריך תגובה 3"]:
                width = 100
            elif col in ["כותב ההערה", "כותב תגובה 1", "כותב תגובה 2", "כותב תגובה 3"]:
                width = 100
            elif col == "הערה":
                width = 250
            elif col in ["תגובה 1", "תגובה 2", "תגובה 3"]:
                width = 250
            else:
                width = 100
                
            # מתיחת עמודות לפי הצורך
            stretch = True if col in ["הערה", "תגובה 1", "תגובה 2", "תגובה 3"] else False
            self.result_tree.column(col, width=width, minwidth=35, anchor="e", stretch=stretch)
        
        # הגדרת סרגלי גלילה
        x_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.result_tree.xview)
        y_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.result_tree.yview)
        self.result_tree.configure(xscrollcommand=x_scrollbar.set, yscrollcommand=y_scrollbar.set)
        
        # סידור רכיבי הטבלה - RTL סדר
        y_scrollbar.pack(side="left", fill="y")  # סרגל גלילה אנכי בצד שמאל
        x_scrollbar.pack(side="bottom", fill="x")  # סרגל הגלילה האופקי למטה
        self.result_tree.pack(side="right", fill="both", expand=True)  # עץ בצד ימין
        
        # טקסט עזרה
        help_frame = ttk.Frame(main_frame)
        help_frame.pack(fill="x", pady=5)
        
        help_text = ttk.Label(help_frame, text="טיפ: ניתן ללחוץ על כותרות העמודות כדי למיין את הטבלה",
                             font=("Segoe UI", 9), foreground="gray")
        help_text.pack(side="right", pady=2)
        
        # מסגרת כפתורים תחתונה
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=5)
        
        # כפתור יציאה
        exit_btn = ttk.Button(button_frame, text="יציאה", command=self.root.quit)
        exit_btn.pack(side="left", padx=5)
        
        # קרדיט
        credit_frame = ttk.Frame(main_frame)
        credit_frame.pack(fill="x", pady=2)
        
        credit = ttk.Label(credit_frame, text="פותח ע\"י יצחק נזרי", 
                          font=("Segoe UI", 9), foreground="gray")
        credit.pack(side="bottom", pady=2)
    
    def select_file(self):
        """בחירת קובץ וורד לעיבוד"""
        self.docx_path = filedialog.askopenfilename(
            filetypes=[("Word Documents", "*.docx")],
            title="בחר קובץ וורד"
        )
        
        if self.docx_path:
            self.file_label.config(text=os.path.basename(self.docx_path))
            self.status_var.set("קובץ נבחר. לחץ על 'עבד את הקובץ' להמשך.")
    
    def process_file(self):
        """עיבוד הקובץ והצגת ההערות בטבלה"""
        if not self.docx_path:
            messagebox.showwarning("שגיאה", "יש לבחור קובץ וורד תחילה")
            return
        
        self.status_var.set("מעבד את הקובץ...")
        self.root.update()
        
        try:
            # ניקוי טבלה קודמת
            self.result_tree.delete(*self.result_tree.get_children())
            self.comments_data = []
            
            # חילוץ שרשורי התגובות - הערה ראשית עם כל התגובות שלה
            comment_threads = self.extract_comment_threads()
            
            if not comment_threads:
                self.status_var.set("לא נמצאו הערות בקובץ.")
                return
            
            # בדיקת ערכי מפתח בהערות לפני הצגה
            keys_to_validate = ['הערה', 'כותב', 'תאריך', 'עמוד']
            for thread in comment_threads:
                for key in keys_to_validate:
                    if key not in thread:
                        thread[key] = ""
            
            # הצגת התוצאות בטבלה (בכיוון RTL - מימין לשמאל)
            for idx, thread in enumerate(comment_threads, 1):
                # הכנת ערכים להצגה בטבלה
                values = [
                    idx,  # מספר
                    thread.get('עמוד', ''),  # עמוד
                    thread.get('הערה', ''),  # הערה מקורית
                    thread.get('כותב', ''),  # כותב הערה
                    thread.get('תאריך', ''),  # תאריך הערה
                ]
                
                # הוספת עד 3 תגובות לשורה
                for i in range(1, 4):
                    values.extend([
                        thread.get(f'תגובה {i}', ''),
                        thread.get(f'כותב תגובה {i}', ''),
                        thread.get(f'תאריך תגובה {i}', '')
                    ])
                
                # הוספה לטבלה - שורות חדשות נוספות לתחילת הטבלה
                # זו דרך נכונה להציג שורות בסדר RTL כדי שהמספרים יהיו הגיוניים
                item_id = self.result_tree.insert("", "end", values=values)
                
                # שמירת האינדקס למידע לשימוש בייצוא
                thread['מס\''] = idx
            
            # שמירת הנתונים במשתנה המחלקה לשימוש בייצוא
            self.comments_data = comment_threads
            
            # עדכון סטטוס
            total_replies = sum(1 for thread in comment_threads for i in range(1, 4) if thread.get(f'תגובה {i}', ''))
            self.status_var.set(f"נמצאו {len(comment_threads)} הערות ו-{total_replies} תגובות. ניתן לייצא לאקסל.")
            
            # מיון הטבלה לפי מספר עמוד אם יש נתונים
            if comment_threads:
                self.sort_treeview_by_column(1)  # מיון לפי עמוד (עמודה שנייה)
            
        except Exception as e:
            error_msg = str(e)
            print(f"שגיאה בעיבוד הקובץ: {error_msg}")
            messagebox.showerror("שגיאה", f"אירעה שגיאה בעיבוד הקובץ:\n{error_msg}")
            self.status_var.set("אירעה שגיאה בעיבוד הקובץ")
            
    def sort_treeview_by_column(self, col, descending=False):
        """מיון הטבלה לפי עמודה מסוימת"""
        try:
            # קבלת נתונים מהטבלה
            data = []
            for child_id in self.result_tree.get_children(''):
                values = self.result_tree.item(child_id, 'values')
                data.append((child_id, values))
            
            # פונקציית מיון לערכים מספריים או טקסטואליים
            def convert_value(val):
                try:
                    # ניסיון להמיר למספר - אם אפשר
                    return int(val)
                except (ValueError, TypeError):
                    # אחרת משאיר כמחרוזת
                    return str(val).lower()
            
            # מיון הנתונים
            data.sort(key=lambda x: convert_value(x[1][col]), reverse=descending)
            
            # סידור מחדש של הטבלה
            for idx, item in enumerate(data):
                self.result_tree.move(item[0], '', idx)
                
            # עדכון מספרי שורה לאחר המיון
            self.update_row_numbers()
            
        except Exception as e:
            print(f"שגיאה במיון הטבלה: {str(e)}")
    
    def update_row_numbers(self):
        """עדכון מספרי שורה בטבלה ובנתונים"""
        try:
            # עדכון מספרי שורה בנתונים
            for idx, child_id in enumerate(self.result_tree.get_children(''), 1):
                # עדכון ערך בטבלה
                values = list(self.result_tree.item(child_id, 'values'))
                values[0] = idx  # עדכון מספר שורה
                self.result_tree.item(child_id, values=values)
                
                # עדכון בנתונים לייצוא
                for thread in self.comments_data:
                    if thread.get('מזהה') == values[5]:  # אם יש התאמה במזהה
                        thread['מס\''] = idx
        except Exception as e:
            print(f"שגיאה בעדכון מספרי שורה: {str(e)}")
    
    def extract_comment_threads(self):
        """
        מחלץ את כל ההערות והתגובות ומארגן אותן בשרשורים
        שרשור = הערה ראשית + כל התגובות שלה
        """
        if not self.docx_path or not os.path.exists(self.docx_path):
            return []
            
        try:
            # ניסיון להשיג מספרי עמודים מדויקים באמצעות Word COM
            page_map = self.get_exact_page_numbers(self.docx_path)
            
            # יצירת אובייקט Document של python-docx
            doc = Document(self.docx_path)
            
            # חישוב מספרי עמודים בשיטה אלטרנטיבית אם אין לנו מידע מדויק
            if not page_map:
                page_map = {}
                self.calculate_page_numbers(doc, page_map)
                
            comment_threads = []  # רשימה סופית של שרשורי הערות
            
            # פתיחת קובץ docx כארכיון ZIP
            with zipfile.ZipFile(self.docx_path, 'r') as docx_zip:
                # חיפוש קובץ XML שמכיל הערות
                comment_files = [f for f in docx_zip.namelist() 
                               if 'word/comments' in f.lower() or 'word/comment' in f.lower()]
                
                if not comment_files:
                    self.status_var.set("לא נמצאו הערות בקובץ")
                    return []
                
                # מיפוי של כל ההערות והתגובות
                all_comments = {}  # מיפוי של כל ההערות/תגובות לפי מזהה
                parent_child_map = {}  # מיפוי של הורים לילדים (הערות לתגובות)
                
                # קריאת כל קבצי ההערות
                for comment_file in comment_files:
                    xml_content = docx_zip.read(comment_file)
                    root = ET.fromstring(xml_content)
                    
                    # זיהוי namespace
                    ns_match = re.search(r'{(.*?)}', root.tag)
                    if not ns_match:
                        continue
                        
                    ns = {'w': ns_match.group(1)}
                    
                    # הדפסת מידע על ההערות למטרות דיבאג
                    self.status_var.set(f"מעבד קובץ הערות: {comment_file}")
                    self.root.update()
                    
                    # חילוץ פרטי ההערות
                    for comment in root.findall('.//w:comment', ns):
                        comment_id = comment.get(f"{{{ns['w']}}}id")
                        if not comment_id:
                            continue
                        
                        # בדיקה אם זו תגובה להערה אחרת
                        parent_id = comment.get(f"{{{ns['w']}}}parentId")
                        
                        # שמירת יחס הורה-ילד
                        if parent_id:
                            if parent_id not in parent_child_map:
                                parent_child_map[parent_id] = []
                            parent_child_map[parent_id].append(comment_id)
                        
                        author = comment.get(f"{{{ns['w']}}}author", "")
                        date = comment.get(f"{{{ns['w']}}}date", "")
                        
                        # חילוץ טקסט ההערה
                        text_parts = []
                        for text_elem in comment.findall('.//w:t', ns):
                            if text_elem.text:
                                text_parts.append(text_elem.text)
                        
                        comment_text = "".join(text_parts)
                        
                        # מספר עמוד
                        page = page_map.get(comment_id, "1")  # ברירת מחדל לעמוד 1
                        
                        # שמירת פרטי ההערה/תגובה
                        all_comments[comment_id] = {
                            'id': comment_id,
                            'parent_id': parent_id,
                            'text': comment_text,
                            'author': author,
                            'date': self.format_date(date),
                            'page': page,
                            'is_reply': parent_id is not None
                        }
                
                # מציאת ההערות הראשיות (שאין להן הורה)
                root_comments = []
                for comment_id, comment in all_comments.items():
                    if not comment['parent_id']:
                        root_comments.append(comment)
                
                # מיון ההערות הראשיות לפי מספר עמוד ואח"כ לפי תאריך
                root_comments.sort(key=lambda x: (int(x['page']), x['date']))
                
                # דיווח על מספר ההערות הראשיות והתגובות
                total_replies = sum(len(children) for children in parent_child_map.values())
                self.status_var.set(f"נמצאו {len(root_comments)} הערות ראשיות ו-{total_replies} תגובות")
                self.root.update()
                
                # בניית שרשורי הערות עם התגובות שלהן
                for root_comment in root_comments:
                    thread = {
                        'מזהה': root_comment['id'],
                        'עמוד': root_comment['page'],
                        'הערה': root_comment['text'],
                        'כותב': root_comment['author'],
                        'תאריך': root_comment['date']
                    }
                    
                    # מציאת כל התגובות של ההערה הנוכחית
                    reply_idx = 1
                    if root_comment['id'] in parent_child_map:
                        # מיון תגובות לפי תאריך
                        replies = [all_comments[reply_id] for reply_id in parent_child_map[root_comment['id']]]
                        replies.sort(key=lambda x: x['date'])
                        
                        # הוספת התגובות לשרשור
                        for reply in replies[:3]:  # מקסימום 3 תגובות
                            thread[f'תגובה {reply_idx}'] = reply['text']
                            thread[f'כותב תגובה {reply_idx}'] = reply['author']
                            thread[f'תאריך תגובה {reply_idx}'] = reply['date']
                            reply_idx += 1
                    
                    # הוספת השרשור לרשימה הסופית
                    comment_threads.append(thread)
                
                return comment_threads
                
        except Exception as e:
            error_message = str(e)
            print(f"שגיאה כללית בחילוץ הערות: {error_message}")
            messagebox.showerror("שגיאה", f"שגיאה בחילוץ הערות:\n{error_message}")
            return []
    
    def get_exact_page_numbers(self, doc_path):
        """מציאת מספרי עמודים מדויקים באמצעות Word COM API"""
        page_map = {}
        
        # אם win32com לא זמין, החזר מילון ריק
        if not self.has_win32com:
            self.status_var.set("חבילת win32com לא מותקנת. משתמש בשיטה חלופית...")
            self.root.update()
            return page_map
            
        try:
            self.status_var.set("מנסה לחשב מספרי עמודים מדויקים באמצעות Word...")
            self.root.update()
            
            # יצירת אובייקט Word
            word_app = win32com.client.Dispatch("Word.Application")
            word_app.Visible = False
            
            # פתיחת המסמך
            abs_path = os.path.abspath(doc_path)
            doc = word_app.Documents.Open(abs_path)
            
            # מספר הערות במסמך
            comment_count = doc.Comments.Count
            self.status_var.set(f"נמצאו {comment_count} הערות. מחשב מספרי עמודים...")
            self.root.update()
            
            # איסוף מזהים אמיתיים של הערות
            comment_ids = []
            try:
                # פתיחת המסמך כ-ZIP להוצאת מזהי הערות מה-XML
                with zipfile.ZipFile(doc_path, 'r') as zip_file:
                    comment_files = [f for f in zip_file.namelist() 
                                   if 'word/comments' in f.lower() or 'word/comment' in f.lower()]
                    
                    for comment_file in comment_files:
                        xml_content = zip_file.read(comment_file)
                        xml_root = ET.fromstring(xml_content)
                        
                        # זיהוי namespace
                        ns_match = re.search(r'{(.*?)}', xml_root.tag)
                        if not ns_match:
                            continue
                            
                        ns = {'w': ns_match.group(1)}
                        
                        # חילוץ מזהי הערות לפי הסדר
                        for comment in xml_root.findall('.//w:comment', ns):
                            comment_id = comment.get(f"{{{ns['w']}}}id")
                            if comment_id:
                                comment_ids.append(comment_id)
            except Exception as e:
                print(f"שגיאה באיסוף מזהי הערות: {str(e)}")
                # אם לא הצלחנו לאסוף מזהי הערות, ניצור מזהים "מלאכותיים"
                comment_ids = [str(i) for i in range(comment_count)]
            
            # מיפוי מספרי הערות למספרי עמודים
            for i in range(1, comment_count + 1):
                try:
                    comment = doc.Comments(i)
                    
                    # מזהה הערה: אם איספנו מזהים, נשתמש בהם; אחרת נשתמש במספר-1
                    comment_id = comment_ids[i-1] if i-1 < len(comment_ids) else str(i-1)
                    
                    # השג את הטקסט עם ההערה
                    comment_range = comment.Scope
                    
                    # מספר העמוד בו נמצאת ההערה (wdActiveEndPageNumber = 3)
                    page_num = comment_range.Information(3)
                    page_map[comment_id] = page_num
                    
                    # הדפסה למטרות דיבאג
                    print(f"הערה {i}, מזהה {comment_id}, עמוד {page_num}")
                    
                except Exception as e:
                    print(f"שגיאה בהערה {i}: {str(e)}")
                    continue
            
            # סגירת המסמך והאפליקציה
            doc.Close(False)
            word_app.Quit()
            
            # הצלחנו לקבל מידע?
            if page_map:
                self.status_var.set(f"חישוב מספרי עמודים הושלם! נמצאו {len(page_map)} הערות.")
                self.root.update()
            
            return page_map
            
        except Exception as e:
            self.status_var.set("לא ניתן להשתמש ב-Word COM API, משתמש בשיטה חלופית...")
            self.root.update()
            print(f"שגיאה בחישוב מספרי עמודים מדויקים: {str(e)}")
            return {}
    
    def calculate_page_numbers(self, doc, page_map):
        """חישוב מספרי עמודים לכל הערה - שיטה משופרת"""
        try:
            self.status_var.set("מחשב מספרי עמודים בשיטה חלופית...")
            self.root.update()
            
            # חישוב הערכה של מספר עמודים לפי תקן מקובל
            chars_per_page = 3000  # הערכה ממוצעת לעמוד טקסט סטנדרטי
            
            # חילוץ כל הפסקאות והמידע עליהן
            paragraphs = []
            para_info = []
            
            # איסוף מידע על כל הפסקאות
            for section in doc.sections:
                for para in doc.paragraphs:
                    paragraphs.append(para)
                    
                    # תכונות של הפסקה
                    para_data = {
                        'text': para.text,
                        'length': len(para.text),
                        'comment_ids': []
                    }
                    
                    # חיפוש מזהי הערות בפסקה
                    for run in para.runs:
                        if hasattr(run, '_element') and run._element is not None:
                            try:
                                comment_references = run._element.xpath(".//w:commentReference")
                                for ref in comment_references:
                                    if ref is not None and ref.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id"):
                                        comment_id = ref.get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id")
                                        para_data['comment_ids'].append(comment_id)
                            except:
                                pass
                    
                    para_info.append(para_data)
            
            # סך כל התווים במסמך
            total_chars = sum(p['length'] for p in para_info)
            
            # הערכת מספר עמודים כולל
            estimated_total_pages = max(3, int(total_chars / chars_per_page) + 1)
            
            # חישוב מיקום יחסי של כל פסקה במסמך
            current_chars = 0
            for i, para_data in enumerate(para_info):
                current_chars += para_data['length']
                
                # מיקום יחסי של הפסקה במסמך (0-1)
                relative_position = current_chars / total_chars if total_chars > 0 else 0
                
                # חישוב מספר עמוד משוער לפי המיקום היחסי
                estimated_page = max(1, min(estimated_total_pages, 
                                          round(relative_position * estimated_total_pages)))
                
                # שמירת מספר העמוד המשוער לכל מזהה הערה בפסקה זו
                for comment_id in para_data['comment_ids']:
                    page_map[comment_id] = estimated_page
            
            # בדיקה עבור מסמכים קצרים - אם יש מעט עמודים, נפזר את ההערות בצורה שווה יותר
            unique_pages = len(set(page_map.values()))
            if unique_pages <= 1 and len(page_map) > 3:
                self.status_var.set("מחשב פיזור הערות משופר...")
                self.root.update()
                
                # מיון מזהי הערות לפי מספר סידורי (מניח שהמזהים הם מספרים)
                comment_ids = sorted(page_map.keys(), key=lambda x: int(x) if x.isdigit() else 0)
                num_comments = len(comment_ids)
                
                # חישוב פיזור משופר
                if estimated_total_pages < 3 and num_comments > 5:
                    # כפה לפחות 3 עמודים למסמכים עם הרבה הערות
                    estimated_total_pages = max(3, num_comments // 5)
                
                # יצירת פיזור הדרגתי על פני העמודים
                for i, comment_id in enumerate(comment_ids):
                    # פיזור הדרגתי על פני העמודים
                    page = max(1, min(estimated_total_pages, 
                                    1 + int((i / max(1, num_comments - 1)) * (estimated_total_pages - 1))))
                    page_map[comment_id] = page
            
            self.status_var.set(f"נמצאו {len(page_map)} הערות בכ-{estimated_total_pages} עמודים")
            self.root.update()
        
        except Exception as e:
            error_msg = str(e)
            print(f"שגיאה בחישוב מספרי עמודים: {error_msg}")
            self.status_var.set("שגיאה בחישוב מספרי עמודים, משתמש בערכי ברירת מחדל")
            self.root.update()
            
            # במקרה של שגיאה - יצירת מספרי עמודים ברירת מחדל
            comment_ids = list(page_map.keys()) if page_map else []
            for i, comment_id in enumerate(comment_ids):
                # התחלה מעמוד 1 ופיזור כל 3-4 הערות
                page_map[comment_id] = (i // 3) + 1
    
    def format_date(self, date_str):
        """עיצוב תאריך לפורמט קריא"""
        if not date_str:
            return ""
        
        try:
            # ניקוי וסטנדרטיזציה של הפורמט
            date_str = date_str.replace('Z', '+00:00')
            date_obj = datetime.fromisoformat(date_str)
            return date_obj.strftime('%d/%m/%Y %H:%M')
        except:
            return date_str
    
    def export_to_excel(self):
        """ייצוא הטבלה לקובץ אקסל עם שמירה על שרשורי התגובות"""
        if not self.comments_data:
            messagebox.showwarning("אזהרה", "אין נתונים לייצוא")
            return
        
        # בחירת מיקום לשמירת הקובץ
        excel_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="שמור קובץ אקסל",
            initialfile=f"הערות_וורד_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        
        if not excel_path:
            return
        
        self.status_var.set("מייצא לאקסל...")
        self.root.update()
        
        try:
            # הכנת הנתונים לאקסל
            excel_data = []
            
            for idx, thread in enumerate(self.comments_data, 1):
                # הוספת מספר שורה
                if 'מס\'' not in thread:
                    thread['מס\''] = idx
                
                row_data = {
                    'מס\'': thread.get('מס\'', ''),
                    'עמוד': thread.get('עמוד', ''),
                    'הערה': thread.get('הערה', ''),
                    'כותב ההערה': thread.get('כותב', ''),
                    'תאריך ההערה': thread.get('תאריך', '')
                }
                
                # הוספת עד 3 תגובות
                for i in range(1, 4):
                    reply_text = thread.get(f'תגובה {i}', '')
                    reply_author = thread.get(f'כותב תגובה {i}', '')
                    reply_date = thread.get(f'תאריך תגובה {i}', '')
                    
                    row_data[f'תגובה {i}'] = reply_text
                    row_data[f'כותב תגובה {i}'] = reply_author
                    row_data[f'תאריך תגובה {i}'] = reply_date
                
                excel_data.append(row_data)
            
            # המרה ל-DataFrame
            df = pd.DataFrame(excel_data)
            
            # הגדרת סדר העמודות - סדר RTL 
            column_order = [
                'מס\'', 'עמוד', 'הערה', 'כותב ההערה', 'תאריך ההערה',
                'תגובה 1', 'כותב תגובה 1', 'תאריך תגובה 1',
                'תגובה 2', 'כותב תגובה 2', 'תאריך תגובה 2',
                'תגובה 3', 'כותב תגובה 3', 'תאריך תגובה 3'
            ]
            
            # מודיע על התקדמות
            self.status_var.set("יוצר קובץ אקסל...")
            self.root.update()
            
            # ייצוא לאקסל עם הגדרות RTL
            try:
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    # יצירת גיליון
                    columns_to_use = [col for col in column_order if col in df.columns]
                    df[columns_to_use].to_excel(writer, index=False, sheet_name='הערות')
                    
                    # הגדרת כיוון RTL לגיליון
                    worksheet = writer.sheets['הערות']
                    worksheet.sheet_view.rightToLeft = True
                    
                    # עיצוב סגנון
                    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                    
                    # הגדרת פונט וסגנון לכותרות
                    header_font = Font(name='Arial', size=11, bold=True)
                    header_fill = PatternFill(start_color="D0E0FF", end_color="D0E0FF", fill_type="solid")
                    
                    # גבולות
                    thin_border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    
                    # עיצוב כותרות
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                        cell.border = thin_border
                    
                    # התאמת רוחב עמודות
                    for i, column in enumerate(columns_to_use):
                        col_idx = i + 1  # אינדקס עמודה באקסל (מתחיל מ-1)
                        
                        # הגדרת רוחב בהתאם לסוג העמודה
                        if column == 'הערה':
                            width = 60  # הערה ראשית רחבה במיוחד
                        elif column in ['תגובה 1', 'תגובה 2', 'תגובה 3']:
                            width = 50  # עמודות תגובה רחבות
                        elif 'תאריך' in column:
                            width = 18  # עמודות תאריך
                        elif 'כותב' in column:
                            width = 20  # עמודות שם כותב
                        elif column == 'עמוד':
                            width = 8  # עמודת מספר עמוד
                        elif column == 'מס\'':
                            width = 6  # עמודת מספור
                        else:
                            width = 20  # ברירת מחדל
                        
                        # התאמה לתוכן בפועל (עד גבול מסוים)
                        try:
                            max_content_length = max(
                                df[column].astype(str).apply(lambda x: len(x) if len(x) < 1000 else 0).max(),
                                len(column)
                            )
                            adjusted_width = min(100, max(width, max_content_length + 2))
                        except:
                            adjusted_width = width
                        
                        # הגדרת רוחב עמודה לפי מיקום
                        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                        
                        # הגדרת יישור לימין עבור כל התאים בעמודה
                        for row_idx in range(2, len(df) + 2):  # מתחיל מהשורה השנייה (אחרי הכותרות)
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                            cell.border = thin_border
                
                # הודעת הצלחה
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה:\n{excel_path}")
                self.status_var.set(f"הנתונים יוצאו בהצלחה: {os.path.basename(excel_path)}")
                
                # פתיחת התיקייה המכילה את הקובץ
                try:
                    os.startfile(os.path.dirname(excel_path))
                except:
                    pass
                    
            except Exception as excel_err:
                # אם נכשל עם openpyxl, ננסה גישה פשוטה יותר
                print(f"שגיאה בכתיבה לאקסל עם openpyxl: {str(excel_err)}")
                self.status_var.set("מנסה שיטת ייצוא חלופית...")
                self.root.update()
                
                # ייצוא בסיסי ללא עיצוב
                df[columns_to_use].to_excel(excel_path, index=False, sheet_name='הערות')
                messagebox.showinfo("הצלחה", f"הקובץ נשמר בהצלחה (ללא עיצוב):\n{excel_path}")
                
        except Exception as e:
            error_msg = str(e)
            print(f"שגיאה בייצוא לאקסל: {error_msg}")
            messagebox.showerror("שגיאה", f"אירעה שגיאה בייצוא לאקסל:\n{error_msg}")
            self.status_var.set("שגיאה בייצוא לאקסל")

def main():
    root = tk.Tk()
    
    # הגדרות RTL ברמת האפליקציה
    try:
        root.tk.call('encoding', 'system', 'utf-8')
        root.tk_strictMotif(False)
    except:
        pass
    
    # בדיקה אם win32com מותקן
    if not HAS_WIN32COM:
        messagebox.showwarning(
            "הערה", 
            "חבילת win32com לא מותקנת. התוכנה תפעל, אבל ללא יכולת לזהות מספרי עמודים מדויקים.\n"
            "להתקנת החבילה הרצו: pip install pywin32"
        )
    
    # יצירת האפליקציה
    app = WordCommentsExtractor(root)
    root.mainloop()

if __name__ == "__main__":
    main()
